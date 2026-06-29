import sqlite3
from unittest.mock import Mock

from langchain_core.runnables import RunnableLambda
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side

from excel_ingestion import (
    BatchAggregateLabels,
    IndexedAggregateLabel,
    extract_sheet_facts,
    ingest_file,
    write_facts,
)

BOLD = Font(bold=True)
TOP_BORDER = Border(top=Side(style="thick"))


def _llm_returning(labels: dict):
    """A fake llm whose `.with_structured_output(...)` returns one is_aggregate verdict per index."""
    structured = RunnableLambda(
        lambda _prompt_value: BatchAggregateLabels(
            labels=[IndexedAggregateLabel(index=i, is_aggregate=v, reasoning="r") for i, v in labels.items()]
        )
    )
    llm = Mock()
    llm.with_structured_output = Mock(return_value=structured)
    return llm


def _build_penjualan_sheet():
    """2-level merged header (Semester > Quarter) starting at row 3, a decorative title row 1,
    and a bold+top-bordered "Jumlah" aggregate row - mirrors sample_data/penjualan_2024.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Penjualan"

    ws.merge_cells("A1:E1")
    ws["A1"] = "Laporan Penjualan Triwulanan 2024"

    ws["A3"] = "Produk"
    ws.merge_cells("B3:C3")
    ws["B3"] = "Semester 1"
    ws.merge_cells("D3:E3")
    ws["D3"] = "Semester 2"
    ws["B4"], ws["C4"], ws["D4"], ws["E4"] = "Q1", "Q2", "Q3", "Q4"

    rows = [("Produk A", 120, 150, 130, 170), ("Produk B", 90, 95, 100, 110)]
    for i, (name, q1, q2, q3, q4) in enumerate(rows):
        r = 5 + i
        ws[f"A{r}"], ws[f"B{r}"], ws[f"C{r}"], ws[f"D{r}"], ws[f"E{r}"] = name, q1, q2, q3, q4

    total_row = 5 + len(rows)
    ws[f"A{total_row}"] = "Jumlah"
    for col, idx in (("B", 1), ("C", 2), ("D", 3), ("E", 4)):
        ws[f"{col}{total_row}"] = sum(row[idx] for row in rows)
    for col in ("A", "B", "C", "D", "E"):
        ws[f"{col}{total_row}"].font = BOLD
        ws[f"{col}{total_row}"].border = TOP_BORDER

    return ws


def _build_cabang_sheet():
    """Single-level header at row 1, quarters in reversed order, aggregate "Total" column at the
    end - mirrors sample_data/penjualan_cabang.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cabang"
    for col, label in zip("ABCDEF", ["Cabang", "Q4", "Q3", "Q2", "Q1", "Total"]):
        ws[f"{col}1"] = label

    branches = [("Jakarta", 200, 190, 180, 170), ("Surabaya", 150, 140, 135, 130)]
    for i, (name, q4, q3, q2, q1) in enumerate(branches):
        r = 2 + i
        ws[f"A{r}"], ws[f"B{r}"], ws[f"C{r}"], ws[f"D{r}"], ws[f"E{r}"] = name, q4, q3, q2, q1
        ws[f"F{r}"] = q4 + q3 + q2 + q1

    return ws


def test_extract_sheet_facts_resolves_multilevel_header_and_flags_aggregate_row():
    ws = _build_penjualan_sheet()

    records, counts = extract_sheet_facts("penjualan_2024.xlsx", "Penjualan", ws)

    by_ref = {r["source_cell_ref"]: r for r in records}
    assert by_ref["Penjualan!B5"]["col_label"] == "Semester 1 > Q1"
    assert by_ref["Penjualan!B5"]["row_label"] == "Produk A"
    assert by_ref["Penjualan!B5"]["is_aggregate"] is False
    assert by_ref["Penjualan!B7"]["row_label"] == "Jumlah"
    assert by_ref["Penjualan!B7"]["is_aggregate"] is True
    assert by_ref["Penjualan!B7"]["value"] == 210.0
    # The decorative title row must not leak into the facts or the header path.
    assert all("Laporan" not in r["col_label"] for r in records)
    assert counts["auto_aggregate"] >= 1
    assert counts["llm_escalated"] == 0


def test_extract_sheet_facts_flags_aggregate_column_regardless_of_position():
    ws = _build_cabang_sheet()

    records, counts = extract_sheet_facts("penjualan_cabang.xlsx", "Cabang", ws)

    by_ref = {r["source_cell_ref"]: r for r in records}
    assert by_ref["Cabang!F2"]["col_label"] == "Total"
    assert by_ref["Cabang!F2"]["is_aggregate"] is True
    assert by_ref["Cabang!F2"]["value"] == 740.0
    # Reversed quarter order must round-trip into col_label, not get reinterpreted as Q1..Q4.
    assert by_ref["Cabang!B2"]["col_label"] == "Q4"
    assert by_ref["Cabang!B2"]["is_aggregate"] is False
    assert counts["llm_escalated"] == 0


def test_extract_sheet_facts_escalates_ambiguous_row_to_llm():
    # "Subtotal" matches the aggregate keyword but its value (99) doesn't actually sum the other
    # rows (10 + 20 = 30) and it has no distinguishing style - this lands in the ambiguous band
    # (score 0.5..0.7) on keyword_match alone, rather than being auto-classified either way.
    wb = Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "Kategori", "Nilai"
    ws["A2"], ws["B2"] = "Item A", 10
    ws["A3"], ws["B3"] = "Item B", 20
    ws["A4"], ws["B4"] = "Subtotal", 99

    llm = _llm_returning({4: True})
    records, counts = extract_sheet_facts("ambiguous.xlsx", "Sheet1", ws, llm=llm)

    by_ref = {r["source_cell_ref"]: r for r in records}
    assert by_ref["Sheet1!B4"]["is_aggregate"] is True
    assert by_ref["Sheet1!B2"]["is_aggregate"] is False
    assert counts["llm_escalated"] == 1
    llm.with_structured_output.assert_called_once()


def test_extract_sheet_facts_defaults_ambiguous_row_to_not_aggregate_without_llm():
    wb = Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "Kategori", "Nilai"
    ws["A2"], ws["B2"] = "Item A", 10
    ws["A3"], ws["B3"] = "Item B", 20
    ws["A4"], ws["B4"] = "Subtotal", 99

    records, counts = extract_sheet_facts("ambiguous.xlsx", "Sheet1", ws, llm=None)

    by_ref = {r["source_cell_ref"]: r for r in records}
    assert by_ref["Sheet1!B4"]["is_aggregate"] is False
    assert counts["defaulted"] == 1
    assert counts["llm_escalated"] == 0


def test_write_facts_reingesting_same_source_file_is_idempotent(tmp_path):
    db_path = tmp_path / "test.db"
    records = [
        {
            "source_file": "f.xlsx", "sheet": "S", "row_label": "A", "col_label": "Q1",
            "value": 1.0, "is_aggregate": False, "source_cell_ref": "S!B2",
        }
    ]

    write_facts(records, db_path)
    write_facts(records, db_path)

    conn = sqlite3.connect(db_path)
    try:
        count = conn.execute("SELECT COUNT(*) FROM excel_facts").fetchone()[0]
    finally:
        conn.close()
    assert count == 1


def test_ingest_file_writes_facts_to_sqlite_and_returns_summary(tmp_path):
    ws = _build_penjualan_sheet()
    xlsx_path = tmp_path / "penjualan_2024.xlsx"
    ws.parent.save(xlsx_path)
    db_path = tmp_path / "test.db"

    summary = ingest_file(xlsx_path, db_path=db_path)

    assert summary.n_facts == 12  # 2 products + 1 total row, x4 quarter columns
    assert summary.auto_aggregate >= 1

    conn = sqlite3.connect(db_path)
    try:
        rows = conn.execute(
            "SELECT row_label, col_label, value, is_aggregate FROM excel_facts WHERE source_cell_ref = 'Penjualan!B7'"
        ).fetchall()
    finally:
        conn.close()
    assert rows == [("Jumlah", "Semester 1 > Q1", 210.0, 1)]
