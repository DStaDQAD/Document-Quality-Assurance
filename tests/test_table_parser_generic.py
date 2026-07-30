import io
from datetime import datetime

import pytest
from openpyxl import Workbook

from table_parser_generic import (
    _load_grid,
    _parse_period,
    _parse_two_row_table,
    parse_generic_grid,
    parse_generic_table,
)


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# _parse_period
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    "raw,expected",
    [
        ("Apr 2026", (2026, "Apr")),
        ("April 2026", (2026, "Apr")),
        ("Apr-26", (2026, "Apr")),
        ("Apr'26", (2026, "Apr")),
        ("Des 2025", (2025, "Dec")),
        ("Mei 2026", (2026, "May")),
        ("2026 Apr", (2026, "Apr")),
        ("2026M04", (2026, "Apr")),
        ("2026-04", (2026, "Apr")),
        ("04/2026", (2026, "Apr")),
        (datetime(2026, 4, 30), (2026, "Apr")),
        ("Q1 2026", (2026, "Q1")),
        ("Q2-2026", (2026, "Q2")),
        ("2026Q2", (2026, "Q2")),
        ("Tw I 2026", (2026, "Q1")),
        ("Tw.II 2026", (2026, "Q2")),
        ("Triwulan II 2026", (2026, "Q2")),
        ("Kuartal IV 2025", (2025, "Q4")),
        ("I 2026", (2026, "Q1")),
        ("2026 III", (2026, "Q3")),
    ],
)
def test_parse_period_recognizes_common_month_header_formats(raw, expected):
    assert _parse_period(raw) == expected


@pytest.mark.parametrize("raw", ["Total", "Harga", "2026", 2026.0, None, "Apr", "I", "Q1", "Q5 2026"])
def test_parse_period_rejects_non_month_headers(raw):
    assert _parse_period(raw) is None


# ---------------------------------------------------------------------------
# Categorical tables (item lists)
# ---------------------------------------------------------------------------

def _build_item_list_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Barang"
    ws.append(["Daftar Barang Elektronik"])
    ws.append(["Nama Barang", "Harga", "Stok"])
    ws.append(["Laptop ASUS", 7_500_000, 10])
    ws.append(["Mouse Logitech", 250_000, 45])
    ws.append(["Keyboard Mechanical", 850_000, 20])
    return _save(wb)


def test_parse_generic_table_detects_categorical_item_list():
    table = parse_generic_table(_build_item_list_bytes(), "Barang")

    assert table.axis_type == "categorical"
    assert table.title == "Daftar Barang Elektronik"
    assert table.row_labels == ["Laptop ASUS", "Mouse Logitech", "Keyboard Mechanical"]
    assert table.col_labels == ["Harga", "Stok"]
    assert table.lookup_cell("Laptop ASUS", "Harga") == 7_500_000.0
    assert table.lookup_cell("Keyboard Mechanical", "Stok") == 20.0


def test_parse_generic_table_skips_leading_numeric_index_column():
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["No", "Nama Barang", "Harga"])
    ws.append([1, "Laptop ASUS", 7_500_000])
    ws.append([2, "Mouse Logitech", 250_000])

    table = parse_generic_table(_save(wb), "S")

    assert table.axis_type == "categorical"
    assert table.row_labels == ["Laptop ASUS", "Mouse Logitech"]
    assert table.col_labels == ["Harga"]
    assert table.lookup_cell("Mouse Logitech", "Harga") == 250_000.0


def test_parse_generic_table_ignores_text_attribute_columns():
    # Text cells (e.g. 'Kategori') cannot be numerically verified — only numeric cells
    # are stored, but the numeric columns around them must still come through.
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Nama Barang", "Kategori", "Harga"])
    ws.append(["Laptop ASUS", "Elektronik", 7_500_000])

    table = parse_generic_table(_save(wb), "S")

    assert table.lookup_cell("Laptop ASUS", "Harga") == 7_500_000.0
    assert table.lookup_cell("Laptop ASUS", "Kategori") is None


# ---------------------------------------------------------------------------
# Temporal tables with combined single-row period headers
# ---------------------------------------------------------------------------

def _build_combined_period_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Penjualan"
    ws.append(["Penjualan per Wilayah"])
    ws.append(["(juta Rp)"])
    ws.append(["Wilayah", "Jan 2026", "Feb 2026", "Mar 2026"])
    ws.append(["Jakarta", 120.0, 130.0, 125.0])
    ws.append(["Surabaya", 80.0, 85.0, 90.0])
    return _save(wb)


def test_parse_generic_table_detects_combined_period_headers_as_temporal():
    table = parse_generic_table(_build_combined_period_bytes(), "Penjualan")

    assert table.axis_type == "temporal"
    assert table.title == "Penjualan per Wilayah"
    assert table.unit == "juta Rp"
    assert table.row_labels == ["Jakarta", "Surabaya"]
    assert table.lookup("Jakarta", 2026, "Jan") == 120.0
    assert table.lookup("Surabaya", 2026, "Mar") == 90.0


def test_parse_generic_table_temporal_lookup_fuzzy_still_works():
    table = parse_generic_table(_build_combined_period_bytes(), "Penjualan")

    matched, value = table.lookup_fuzzy("jakarta", 2026, "Feb")

    assert matched == "Jakarta"
    assert value == 130.0


def test_parse_generic_table_reads_excel_date_cells_as_periods():
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Metrik", datetime(2026, 1, 31), datetime(2026, 2, 28)])
    ws.append(["Produksi", 500.0, 520.0])

    table = parse_generic_table(_save(wb), "S")

    assert table.axis_type == "temporal"
    assert table.lookup("Produksi", 2026, "Jan") == 500.0
    assert table.lookup("Produksi", 2026, "Feb") == 520.0


# ---------------------------------------------------------------------------
# Two-row header path (year row merged over quarter tokens — BI survey layout)
# ---------------------------------------------------------------------------

def _build_quarterly_survey_bytes():
    """Mirror the BI survey Tabel1 shape: title, ID+EN label block, years merged over
    4 quarter columns, vertical category merges, an unmerged blank label cell, an 'na'
    sentinel, a Total row with a blank child column, and a merged footnote row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Survei"
    # Row 1 blank; row 2 title in column B.
    ws["B2"] = "Tabel 1. Penyaluran Kredit Baru"
    # Row 3: label-block headers + int years merged across their quarters.
    ws["B3"], ws["C3"] = "Jenis Kredit", "Rincian Kredit"
    ws["D3"], ws["E3"] = "Type of Loans", "Loans in Detail"
    ws["F3"] = 2012
    ws.merge_cells("F3:I3")
    ws["J3"] = 2013
    ws.merge_cells("J3:M3")
    # Row 4: quarter tokens.
    for col, tok in zip("FGHIJKLM", ["I", "II", "III", "IV"] * 2):
        ws[f"{col}4"] = tok
    # Data rows.
    ws["B5"], ws["D5"] = "Menurut Penggunaan", "Based on Usage"
    ws.merge_cells("B5:B7")
    ws.merge_cells("D5:D7")
    ws["C5"], ws["E5"] = "Kredit Modal Kerja", "Working Capital Loans"
    for col, v in zip("FGHIJKLM", [63.55, 92.11, 82.45, 73.42, 33.74, 70.47, 80.9, 76.09]):
        ws[f"{col}5"] = v
    ws["C6"], ws["E6"] = "Kredit Investasi", "Investment Loans"
    for col, v in zip("FGHIJKLM", [73.98, 88.38, 78.09, 83.75, 53.49, 82.02, 58.53, 45.69]):
        ws[f"{col}6"] = v
    ws["C7"], ws["E7"] = "Kredit Konsumsi", "Consumer Loans"
    ws["F7"] = "na"  # missing-data sentinel — must not become a stored key
    for col, v in zip("GHIJKLM", [43.61, 19.1, 49.52, 6.85, 31.76, 24.98, 14.64]):
        ws[f"{col}7"] = v
    # Unmerged category label with blank cells below it (row 9 must inherit it).
    ws["B8"], ws["C8"], ws["E8"] = "Sektor Ekonomi", "Pertanian", "Agriculture"
    for col, v in zip("FGHIJKLM", [28.3, 57.26, 40.29, 66.5, 37.73, 65.38, 49.98, 47.73]):
        ws[f"{col}8"] = v
    ws["C9"], ws["E9"] = "Perikanan", "Fishery"
    for col, v in zip("FGHIJKLM", [22.6, 37.72, 25.27, 45.34, 2.94, 39.49, 25.87, 34.12]):
        ws[f"{col}9"] = v
    # Total row: child column C blank — label must stay exactly 'Total'.
    ws["B10"] = "Total"
    for col, v in zip("FGHIJKLM", [50.0, 60.0, 55.0, 58.0, 40.0, 61.0, 52.0, 49.0]):
        ws[f"{col}10"] = v
    # Footnote merged across the full width.
    ws["B11"] = "*) angka sementara"
    ws.merge_cells("B11:M11")
    return _save(wb)


_KMK_LABEL = "Menurut Penggunaan > Kredit Modal Kerja > Based on Usage > Working Capital Loans"


def test_two_row_quarter_table_parses_temporal():
    table = parse_generic_table(_build_quarterly_survey_bytes(), "Survei")

    assert table.axis_type == "temporal"
    assert table.title == "Tabel 1. Penyaluran Kredit Baru"
    assert table.lookup(_KMK_LABEL, 2012, "Q1") == 63.55
    assert table.lookup(_KMK_LABEL, 2013, "Q2") == 70.47


def test_two_row_na_sentinel_is_not_stored():
    table = parse_generic_table(_build_quarterly_survey_bytes(), "Survei")
    konsumsi = next(l for l in table.row_labels if "Kredit Konsumsi" in l)

    assert table.lookup(konsumsi, 2012, "Q1") is None
    assert table.lookup(konsumsi, 2012, "Q2") == 43.61


def test_two_row_composite_labels_and_fuzzy_lookup():
    table = parse_generic_table(_build_quarterly_survey_bytes(), "Survei")

    label, value = table.lookup_fuzzy("Kredit Modal Kerja", 2013, "Q2")
    assert label == _KMK_LABEL
    assert value == 70.47


def test_two_row_blank_label_inherits_category_above():
    table = parse_generic_table(_build_quarterly_survey_bytes(), "Survei")

    assert table.lookup("Sektor Ekonomi > Perikanan > Fishery", 2012, "Q2") == 37.72


def test_two_row_total_row_does_not_inherit_child_qualifier():
    table = parse_generic_table(_build_quarterly_survey_bytes(), "Survei")

    assert "Total" in table.row_labels
    assert table.lookup("Total", 2013, "Q4") == 49.0


def test_two_row_footnote_row_is_excluded():
    table = parse_generic_table(_build_quarterly_survey_bytes(), "Survei")

    assert not any("angka sementara" in l for l in table.row_labels)


def _build_stacked_blocks_bytes():
    """Two stacked tables in one sheet (the Tabel3 shape): the header block repeats
    after a blank row, with a merged sub-header row between quarters and data, and
    the leftmost 'Periode' label column disambiguating the segments."""
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["B1"] = "Tabel 3. Prakiraan"
    # Block 1.
    ws["B2"], ws["C2"] = "Periode", "Jenis Simpanan"
    ws["D2"] = 2012
    ws.merge_cells("D2:G2")
    ws["H2"] = 2013
    ws.merge_cells("H2:K2")
    for col, tok in zip("DEFGHIJK", ["I", "II", "III", "IV"] * 2):
        ws[f"{col}3"] = tok
    ws["B4"], ws["C4"] = "Prakiraan per Triwulan", "Giro"
    for col, v in zip("DEFGHIJK", [10.0, 11.0, 12.0, 13.0, 14.0, 91.9, 16.0, 17.0]):
        ws[f"{col}4"] = v
    ws["C5"] = "Tabungan"
    for col, v in zip("DEFGHIJK", [20.0, 21.0, 22.0, 23.0, 24.0, 25.0, 26.0, 27.0]):
        ws[f"{col}5"] = v
    # Row 6 blank, then block 2.
    ws["B7"], ws["C7"] = "Periode", "Jenis Simpanan"
    ws["D7"] = 2012
    ws.merge_cells("D7:G7")
    ws["H7"] = 2013
    ws.merge_cells("H7:K7")
    for col, tok in zip("DEFGHIJK", ["I", "II", "III", "IV"] * 2):
        ws[f"{col}8"] = tok
    # Sub-header row merged across the data columns (must not become a data row).
    ws["D9"] = "Prakiraan / Estimation"
    ws.merge_cells("D9:K9")
    ws["B10"], ws["C10"] = "Prakiraan Selama Setahun", "Giro"
    for col, v in zip("DEFGHIJK", [30.0, 31.0, 32.0, 33.0, 77.2, 35.0, 36.0, 37.0]):
        ws[f"{col}10"] = v
    return _save(wb)


def test_two_row_stacked_blocks_are_disambiguated_by_leftmost_label():
    table = parse_generic_table(_build_stacked_blocks_bytes(), "S")

    assert table.lookup("Prakiraan per Triwulan > Giro", 2013, "Q2") == 91.9
    assert table.lookup("Prakiraan Selama Setahun > Giro", 2013, "Q1") == 77.2
    assert not any("Estimation" in l for l in table.row_labels)


def test_two_row_declines_sparse_year_month_layout():
    # Sparse single-cell year anchors above a month row = the BI monthly layout,
    # which tier 1 owns; the two-row path must decline (contiguous-run guard).
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "Uraian"
    ws["B1"] = 2025
    ws["F1"] = 2026
    for col, m in zip("BCDE", ["Jan", "Feb", "Mar", "Apr"]):
        ws[f"{col}2"] = m
    for col, m in zip("FGHI", ["Jan", "Feb", "Mar", "Apr"]):
        ws[f"{col}2"] = m
    ws["A3"] = "Uang Beredar"
    for col, v in zip("BCDEFGHI", [1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0]):
        ws[f"{col}3"] = v

    assert _parse_two_row_table(_load_grid(_save(wb), "S")) is None


def test_two_row_accepts_merged_year_runs_with_month_tokens():
    # Merged (contiguous) years over month tokens IS this layout — accepted.
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "Uraian"
    ws["B1"] = 2026
    ws.merge_cells("B1:C1")
    ws["D1"] = 2027
    ws.merge_cells("D1:E1")
    for col, m in zip("BCDE", ["Jan", "Feb", "Jan", "Feb"]):
        ws[f"{col}2"] = m
    ws["A3"] = "Uang Beredar"
    for col, v in zip("BCDE", [1.0, 2.0, 3.0, 4.0]):
        ws[f"{col}3"] = v

    table = parse_generic_table(_save(wb), "S")

    assert table.axis_type == "temporal"
    assert table.lookup("Uang Beredar", 2027, "Feb") == 4.0


def test_single_row_categorical_column_named_I_unchanged():
    # A categorical column literally named "I" (no year row above) must stay categorical.
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Nama", "I", "II"])
    ws.append(["A", 1, 2])
    ws.append(["B", 3, 4])

    table = parse_generic_table(_save(wb), "S")

    assert table.axis_type == "categorical"
    assert table.lookup_cell("A", "I") == 1.0


def test_single_row_combined_quarter_headers_become_temporal():
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Produk", "Q1 2026", "Q2 2026"])
    ws.append(["Sepatu", 100, 120])

    table = parse_generic_table(_save(wb), "S")

    assert table.axis_type == "temporal"
    assert table.lookup("Sepatu", 2026, "Q2") == 120.0


# ---------------------------------------------------------------------------
# Merged-cell forward fill in _load_grid
# ---------------------------------------------------------------------------

def test_load_grid_scales_percent_formatted_cells_to_display_value():
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = 0.0577
    ws["A1"].number_format = "0.00%"
    ws["B1"] = 0.0577          # unformatted — must stay raw
    ws["C1"] = 42
    ws["C1"].number_format = '0"%"'  # quoted literal % — Excel does not scale these

    grid = _load_grid(_save(wb), "S")

    assert grid[0][0] == pytest.approx(5.77)
    assert grid[0][1] == pytest.approx(0.0577)
    assert grid[0][2] == 42


def test_load_grid_forward_fills_merged_cells():
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["B2"] = 2026                      # horizontal year merge across 4 columns
    ws.merge_cells("B2:E2")
    ws["A4"] = "Sektor Ekonomi"          # vertical category merge across 3 rows
    ws.merge_cells("A4:A6")

    grid = _load_grid(_save(wb), "S")

    assert [grid[1][c] for c in range(1, 5)] == [2026, 2026, 2026, 2026]
    assert [grid[r][0] for r in range(3, 6)] == ["Sektor Ekonomi"] * 3


# ---------------------------------------------------------------------------
# Failure modes
# ---------------------------------------------------------------------------

def test_parse_generic_table_raises_on_unrecognized_file_format():
    with pytest.raises(ValueError, match="Unrecognized file format"):
        parse_generic_table(b"not an excel file at all", "S")


def test_parse_generic_table_raises_when_sheet_not_found():
    with pytest.raises(ValueError, match="not found"):
        parse_generic_table(_build_item_list_bytes(), "WrongSheet")


def test_parse_generic_table_raises_when_no_header_structure_found():
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    for _ in range(5):
        ws.append(["hanya teks tanpa angka sama sekali"])

    with pytest.raises(ValueError, match="header"):
        parse_generic_table(_save(wb), "S")


# ---------------------------------------------------------------------------
# Hierarchical label blocks: bullet/enumeration markers and redundant qualifiers
# ---------------------------------------------------------------------------

def _build_sectioned_survey_bytes():
    """Mirror the consumer-survey sheet shape: an enumeration marker column ('A.', 'B1.'),
    a bullet column ('-') that shares its column with the SECTION NAME, and the section's
    member rows below. Two sections repeat the same member names, one does not."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SK"
    ws["C1"] = "Tabel 2. Indeks per Kelompok Pengeluaran"
    ws["C3"] = "KETERANGAN"
    ws["G3"] = 2026
    ws.merge_cells("G3:J3")
    for col, tok in zip("GHIJ", ["Jan", "Feb", "Mar", "Apr"]):
        ws[f"{col}4"] = tok
    # Section A — header row carries no data; its name sits in the same column as the
    # bullets of the rows below it.
    ws["D5"], ws["E5"] = "A.", "Indeks Keyakinan Konsumen (IKK)"
    ws["E6"], ws["F6"] = "- ", "Pengeluaran Rp1 - 2 juta"
    for col, v in zip("GHIJ", [101.0, 102.0, 103.0, 104.0]):
        ws[f"{col}6"] = v
    ws["E7"], ws["F7"] = "- ", "Pengeluaran >Rp5 juta"
    for col, v in zip("GHIJ", [121.1, 121.2, 121.3, 121.4]):
        ws[f"{col}7"] = v
    # Section B repeats the same member names.
    ws["D8"], ws["E8"] = "B.", "Indeks Kondisi Ekonomi (IKE)"
    ws["E9"], ws["F9"] = "- ", "Pengeluaran Rp1 - 2 juta"
    for col, v in zip("GHIJ", [91.0, 92.0, 93.0, 94.0]):
        ws[f"{col}9"] = v
    ws["E10"], ws["F10"] = "- ", "Pengeluaran >Rp5 juta"
    for col, v in zip("GHIJ", [111.1, 111.2, 111.3, 111.4]):
        ws[f"{col}10"] = v
    # Section C's single member name is unique in the sheet.
    ws["D11"], ws["E11"] = "C.", "Indeks Ekspektasi Konsumen (IEK)"
    ws["E12"], ws["F12"] = "- ", "Rencana Pembelian Rumah"
    for col, v in zip("GHIJ", [51.0, 52.0, 53.0, 54.0]):
        ws[f"{col}12"] = v
    return _save(wb)


def test_two_row_bullet_marker_does_not_replace_the_section_name():
    # The '-' bullet shares its column with the section title; treating it as a real label
    # buried the title and left every section's rows labelled identically.
    table = parse_generic_table(_build_sectioned_survey_bytes(), "SK")

    assert "Indeks Keyakinan Konsumen (IKK) > Pengeluaran >Rp5 juta" in table.row_labels
    assert "Indeks Kondisi Ekonomi (IKE) > Pengeluaran >Rp5 juta" in table.row_labels
    assert not any(" - > " in label for label in table.row_labels)


def test_two_row_enumeration_marker_is_not_part_of_the_label():
    table = parse_generic_table(_build_sectioned_survey_bytes(), "SK")

    assert not any(label.startswith(("A.", "B.", "C.")) for label in table.row_labels)


def test_two_row_repeated_member_names_stay_qualified_per_section():
    table = parse_generic_table(_build_sectioned_survey_bytes(), "SK")

    assert table.lookup("Indeks Keyakinan Konsumen (IKK) > Pengeluaran >Rp5 juta", 2026, "Apr") == 121.4
    assert table.lookup("Indeks Kondisi Ekonomi (IKE) > Pengeluaran >Rp5 juta", 2026, "Apr") == 111.4


def test_two_row_unique_member_drops_its_section_header_qualifier():
    # 'Rencana Pembelian Rumah' occurs once, so its section header adds no information —
    # and those borrowed words would make the row look farther from a claim naming it.
    table = parse_generic_table(_build_sectioned_survey_bytes(), "SK")

    assert "Rencana Pembelian Rumah" in table.row_labels
    assert table.lookup("Rencana Pembelian Rumah", 2026, "Apr") == 54.0


def test_two_row_qualifier_from_a_data_row_is_never_dropped():
    # 'Sektor Ekonomi' is carried from a row that HAS data (Pertanian's), so it is part of
    # that block's identity — unlike a header-only section title. The bilingual label block
    # of the same fixture (all parts explicit) must survive intact too.
    table = parse_generic_table(_build_quarterly_survey_bytes(), "Survei")

    assert "Sektor Ekonomi > Perikanan > Fishery" in table.row_labels
    assert _KMK_LABEL in table.row_labels


# ---------------------------------------------------------------------------
# The grid seam: parse_generic_table is now a thin wrapper over parse_generic_grid, so a
# non-spreadsheet source (tables transcribed out of a PDF) can reuse the same heuristics.
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("builder,sheet", [
    (_build_item_list_bytes, "Barang"),
    (_build_combined_period_bytes, "Penjualan"),
    (_build_quarterly_survey_bytes, "Survei"),
    (_build_stacked_blocks_bytes, "S"),
    (_build_sectioned_survey_bytes, "SK"),
])
def test_parse_generic_grid_matches_the_bytes_wrapper(builder, sheet):
    data = builder()
    assert parse_generic_grid(_load_grid(data, sheet)) == parse_generic_table(data, sheet)
