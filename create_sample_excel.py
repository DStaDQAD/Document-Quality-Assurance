"""
Generates synthetic "stylized" Excel files into `sample_data/`, used as development/test
fixtures for `excel_ingestion.py` until real source files are available.

Run with:
    python create_sample_excel.py

Two files are generated, deliberately shaped differently from each other:

- penjualan_2024.xlsx: a decorative merged title row, a 2-level merged header (Semester > Q1..Q4),
  row labels starting at column A but the header band starting at row 3 (not row 1), and a bold
  "Jumlah" total row at the bottom with a top border.
- penjualan_cabang.xlsx: a single-level header starting at row 1 (no offset, no merge), quarter
  columns in reversed order (Q4..Q1), and an aggregate "Total" column instead of an aggregate row.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SAMPLE_DIR = Path(__file__).parent / "sample_data"

BOLD = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOP_BORDER = Border(top=Side(style="thick"))
CENTER = Alignment(horizontal="center")


def _build_penjualan_2024() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Penjualan"

    # Decorative title row spanning the whole table width - not a header, not data.
    ws.merge_cells("A1:E1")
    ws["A1"] = "Laporan Penjualan Triwulanan 2024"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = CENTER
    # Row 2 left blank on purpose (spacer row between title and header band).

    # Row 3-4: 2-level merged header. Header band starts at row 3, not row 1.
    ws["A3"] = "Produk"
    ws["A3"].font = BOLD
    ws.merge_cells("B3:C3")
    ws["B3"] = "Semester 1"
    ws.merge_cells("D3:E3")
    ws["D3"] = "Semester 2"
    for col in ("A", "B", "D"):
        ws[f"{col}3"].font = BOLD
        ws[f"{col}3"].fill = HEADER_FILL
        ws[f"{col}3"].alignment = CENTER

    sub_headers = {"B4": "Q1", "C4": "Q2", "D4": "Q3", "E4": "Q4"}
    for ref, label in sub_headers.items():
        ws[ref] = label
        ws[ref].font = BOLD
        ws[ref].fill = HEADER_FILL
        ws[ref].alignment = CENTER

    # Data rows.
    data_rows = [
        ("Produk A", 120, 150, 130, 170),
        ("Produk B", 90, 95, 100, 110),
        ("Produk C", 60, 70, 75, 80),
    ]
    for i, (name, q1, q2, q3, q4) in enumerate(data_rows):
        r = 5 + i
        ws[f"A{r}"] = name
        ws[f"B{r}"] = q1
        ws[f"C{r}"] = q2
        ws[f"D{r}"] = q3
        ws[f"E{r}"] = q4

    # Aggregate row: bold "Jumlah" label + bold sums + top border, sitting right under the data.
    totals_row = 5 + len(data_rows)
    ws[f"A{totals_row}"] = "Jumlah"
    for col_letter, idx in (("B", 1), ("C", 2), ("D", 3), ("E", 4)):
        total = sum(row[idx] for row in data_rows)
        ws[f"{col_letter}{totals_row}"] = total
    for col_letter in ("A", "B", "C", "D", "E"):
        cell = ws[f"{col_letter}{totals_row}"]
        cell.font = BOLD
        cell.border = TOP_BORDER

    for col_letter, width in (("A", 14), ("B", 10), ("C", 10), ("D", 10), ("E", 10)):
        ws.column_dimensions[col_letter].width = width

    return wb


def _build_penjualan_cabang() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cabang"

    # Single-level header starting right at row 1, column A - no offset, no merge.
    # Quarters deliberately reversed (Q4..Q1) and an aggregate "Total" column appended at the end,
    # to prove the parser doesn't assume a fixed column order or aggregate-as-last-row shape.
    headers = ["Cabang", "Q4", "Q3", "Q2", "Q1", "Total"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    # branch, Q4, Q3, Q2, Q1 (Total computed as the sum of the four quarter values).
    branches = [
        ("Jakarta", 200, 190, 180, 170),
        ("Surabaya", 150, 140, 135, 130),
        ("Bandung", 110, 105, 100, 95),
    ]
    for i, (name, q4, q3, q2, q1) in enumerate(branches):
        r = 2 + i
        values = [name, q4, q3, q2, q1, q4 + q3 + q2 + q1]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=r, column=col_idx, value=value)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    return wb


def main() -> None:
    SAMPLE_DIR.mkdir(exist_ok=True)

    targets = {
        "penjualan_2024.xlsx": _build_penjualan_2024,
        "penjualan_cabang.xlsx": _build_penjualan_cabang,
    }
    for filename, builder in targets.items():
        path = SAMPLE_DIR / filename
        builder().save(path)
        print(f"Wrote {path}")


if __name__ == "__main__":
    main()
