"""Heuristic parser for arbitrary 'clean' tables — Tier 2 of the parsing cascade.

parse_bi_table (Tier 1) handles the known BI wide time-series layout. This module handles
everything else with a detectable header structure:

  - Two-row headers: an int-year row merged across period columns above a row of bare
    quarter ("I/II/III/IV", "Q1", "Tw I") or month tokens — the BI survey layout, incl.
    several tables stacked in one sheet and multi-column hierarchical label blocks.
  - Time-series tables whose period headers are COMBINED in one row ("Apr 2026", "2026M04",
    "Apr-26", "Q1 2026", "Triwulan II 2026", real Excel date cells).
  - Non-time-series (categorical) tables such as item lists / inventories / budgets:
    a header row of attribute names ("Nama Barang | Harga | Stok") over data rows.

Quarter columns share the temporal key space with months: TableData keys are
(label, year, "Apr") or (label, year, "Q2").

Detection strategy (all deterministic — no LLM, values never leave code):
  1. Header row  : first row (within the top 20) with >= 2 non-empty cells where >= 80% of
                   the non-empty cells are text or period-parseable, and at least one of the
                   next 6 rows carries a numeric cell.
  2. Label column: leftmost column whose cells below the header are >= 60% non-numeric text.
                   Columns left of it (e.g. numeric row-index columns) are ignored.
  3. Axis kind   : if >= 60% of the data-column headers (min 2) parse as (year, month)
                   periods, the table is TEMPORAL and is stored under (label, year, month)
                   keys; otherwise it is CATEGORICAL and stored under (label, col_name) keys.
  4. Title/unit  : non-empty rows above the header; a row fully wrapped in parentheses is the
                   unit, the first other row is the title.

Known limitations (deliberate Tier-2 scope — messier layouts belong to the future LLM
structure-mapping tier):
  - Only numeric cell values are stored; text attributes (e.g. 'Kategori') are not
    verifiable and are skipped.
  - The leftmost text column is assumed to be the row key. If an ID/code column precedes
    the human-readable name column, rows are keyed by the code.
  - Duplicate row labels / column headers keep the first occurrence.
  - The two-row path requires merged-style contiguous year runs for month tokens, so the
    sparse-year-anchor BI monthly layout still belongs to Tier 1 exclusively.
"""

import io
import re
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

from table_model import QUAL_SEP, TableData

_MONTH_ABBREVS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Canonical quarter tokens. They share the (year, token) period-key slot with month
# abbreviations: TableData's temporal keys are (label, year, "Apr") or (label, year, "Q2").
_QUARTER_CANON = ["Q1", "Q2", "Q3", "Q4"]

# Quarter spellings (lowercased) -> canonical token. Covers roman numerals as used in BI
# survey headers ('I'..'IV'), Q-forms, and Indonesian 'Tw'/'Triwulan'/'Kuartal' prefixes.
_QUARTER_TOKENS: Dict[str, str] = {}
for _i, _q in enumerate(_QUARTER_CANON, start=1):
    _roman = ["i", "ii", "iii", "iv"][_i - 1]
    for _form in (_roman, f"q{_i}", f"q {_i}", f"tw{_i}", f"tw {_i}", f"tw {_roman}",
                  f"triwulan {_i}", f"triwulan {_roman}",
                  f"kuartal {_i}", f"kuartal {_roman}"):
        _QUARTER_TOKENS[_form] = _q

# Indonesian + English month names/abbreviations -> canonical 3-letter English abbreviation
# (the same canonical form structured_extractor emits and TableData keys on).
_MONTH_NAMES: Dict[str, str] = {
    "jan": "Jan", "januari": "Jan", "january": "Jan",
    "feb": "Feb", "februari": "Feb", "february": "Feb", "peb": "Feb",
    "mar": "Mar", "mrt": "Mar", "maret": "Mar", "march": "Mar",
    "apr": "Apr", "april": "Apr",
    "may": "May", "mei": "May",
    "jun": "Jun", "juni": "Jun", "june": "Jun",
    "jul": "Jul", "juli": "Jul", "july": "Jul",
    "aug": "Aug", "agu": "Aug", "ags": "Aug", "agt": "Aug", "agustus": "Aug", "august": "Aug",
    "sep": "Sep", "sept": "Sep", "september": "Sep",
    "oct": "Oct", "okt": "Oct", "oktober": "Oct", "october": "Oct",
    "nov": "Nov", "nop": "Nov", "november": "Nov",
    "dec": "Dec", "des": "Dec", "desember": "Dec", "december": "Dec",
}

_MAX_HEADER_SCAN = 20   # rows from the top considered as header candidates
_DATA_PROBE_ROWS = 6    # rows below a header candidate probed for numeric data
_HEADERISH_MIN_FRAC = 0.8
_LABEL_TEXT_MIN_FRAC = 0.6
_PERIOD_MIN_FRAC = 0.6

_NAME_YEAR_RE = re.compile(r"^([A-Za-z]+)[\s.\-/']*(\d{2}|\d{4})$")
_YEAR_NAME_RE = re.compile(r"^(\d{4})[\s.\-/']*([A-Za-z]+)$")
_YEAR_M_NUM_RE = re.compile(r"^(\d{4})[Mm\-/](\d{1,2})$")
_NUM_YEAR_RE = re.compile(r"^(\d{1,2})[\-/](\d{4})$")

# Combined quarter+year single-cell forms: "Q1 2026", "Tw I 2026", "Triwulan II 2026",
# "2026Q1", "I 2026". Bare tokens without a year are handled by _parse_quarter_token
# (two-row header path) instead, so a lone categorical column named "I" stays categorical.
_QTR_PHRASE = r"(?:q|tw|triwulan|kuartal)[\s.]*(?:[1-4]|[ivIV]+)|[ivIV]+"
_QTR_YEAR_RE = re.compile(rf"^({_QTR_PHRASE})[\s.\-/']*(\d{{2}}|\d{{4}})$", re.IGNORECASE)
_YEAR_QTR_RE = re.compile(rf"^(\d{{4}})[\s.\-/']*({_QTR_PHRASE})$", re.IGNORECASE)


def _normalize_quarter_phrase(raw: str) -> Optional[str]:
    """Canonicalize a quarter phrase ('Tw.II', 'q1', 'IV') to 'Q1'..'Q4', else None."""
    return _QUARTER_TOKENS.get(re.sub(r"[\s.]+", " ", raw).strip().lower())


def _parse_quarter_token(value) -> Optional[str]:
    """Return 'Q1'..'Q4' when a header cell denotes one bare calendar quarter, else None.

    Used only by the two-row header path, where the year comes from the row above.
    """
    if not isinstance(value, str):
        return None
    text = value.strip().rstrip("*").strip()
    if not text:
        return None
    return _normalize_quarter_phrase(text)


def _normalize_year(raw: str) -> Optional[int]:
    year = int(raw)
    if len(raw) == 2:
        year += 2000 if year < 70 else 1900
    return year if 1990 <= year <= 2100 else None


def _parse_period(value) -> Optional[Tuple[int, str]]:
    """Return (year, month_abbrev) when a header cell denotes one calendar month, else None."""
    if isinstance(value, (datetime, date)):
        if 1990 <= value.year <= 2100:
            return value.year, _MONTH_ABBREVS[value.month - 1]
        return None
    if not isinstance(value, str):
        return None
    text = re.sub(r"\s+", " ", value).strip().rstrip("*").strip()
    if not text:
        return None

    m = _NAME_YEAR_RE.match(text)
    if m:
        month = _MONTH_NAMES.get(m.group(1).lower())
        year = _normalize_year(m.group(2))
        if month and year:
            return year, month
        # Not a month name — a roman quarter ("I 2026") also fits [A-Za-z]+.
        quarter = _QUARTER_TOKENS.get(m.group(1).lower())
        if quarter and year:
            return year, quarter
        return None
    m = _YEAR_NAME_RE.match(text)
    if m:
        month = _MONTH_NAMES.get(m.group(2).lower())
        year = _normalize_year(m.group(1))
        if month and year:
            return year, month
        quarter = _QUARTER_TOKENS.get(m.group(2).lower())
        if quarter and year:
            return year, quarter
        return None
    m = _YEAR_M_NUM_RE.match(text)
    if m:
        year, month_num = _normalize_year(m.group(1)), int(m.group(2))
        if year and 1 <= month_num <= 12:
            return year, _MONTH_ABBREVS[month_num - 1]
        return None
    m = _NUM_YEAR_RE.match(text)
    if m:
        month_num, year = int(m.group(1)), _normalize_year(m.group(2))
        if year and 1 <= month_num <= 12:
            return year, _MONTH_ABBREVS[month_num - 1]
        return None
    m = _QTR_YEAR_RE.match(text)
    if m:
        quarter = _normalize_quarter_phrase(m.group(1))
        year = _normalize_year(m.group(2))
        if quarter and year:
            return year, quarter
        return None
    m = _YEAR_QTR_RE.match(text)
    if m:
        quarter = _normalize_quarter_phrase(m.group(2))
        year = _normalize_year(m.group(1))
        if quarter and year:
            return year, quarter
    return None


# ---------------------------------------------------------------------------
# Grid loading (.xls via xlrd, .xlsx via openpyxl) — values only, dates as datetime.
# Merged cells are forward-filled: the anchor (top-left) value is copied into every
# cell of the merge, so merged year headers and vertically-merged category labels
# become visible to the structure heuristics.
# ---------------------------------------------------------------------------

def _apply_merged_fill(grid: List[List], ranges: List[Tuple[int, int, int, int]]) -> None:
    """Copy each merged range's anchor value into all its cells, in place.

    `ranges` uses xlrd's convention: 0-based half-open (rlo, rhi, clo, chi).
    Rows shorter than a range's right edge are extended with None first.
    """
    for rlo, rhi, clo, chi in ranges:
        if rlo >= len(grid):
            continue
        anchor = grid[rlo][clo] if clo < len(grid[rlo]) else None
        if anchor is None:
            continue
        for r in range(rlo, min(rhi, len(grid))):
            row = grid[r]
            if len(row) < chi:
                row.extend([None] * (chi - len(row)))
            for c in range(clo, chi):
                row[c] = anchor


def _load_grid(data: bytes, sheet_name: str) -> List[List]:
    if data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        import xlrd

        # formatting_info exposes merged_cells; some .xls files can't provide it —
        # fall back to a plain open and skip the merged fill (same pattern as
        # excel_parser_bi._read_xls).
        try:
            wb = xlrd.open_workbook(file_contents=data, formatting_info=True)
        except Exception:
            wb = xlrd.open_workbook(file_contents=data)
        try:
            ws = wb.sheet_by_name(sheet_name)
        except xlrd.XLRDError:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheet_names()}")
        grid: List[List] = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                v = ws.cell_value(r, c)
                if ws.cell_type(r, c) == xlrd.XL_CELL_DATE:
                    v = xlrd.xldate.xldate_as_datetime(v, wb.datemode)
                row.append(v)
            grid.append(row)
        _apply_merged_fill(grid, list(getattr(ws, "merged_cells", []) or []))
        return grid
    elif data[:4] == b'PK\x03\x04':
        import openpyxl

        # Not read_only: merged_cells.ranges is unavailable in read-only mode, and the
        # workbooks handled here are small enough to load fully.
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        try:
            ws = wb[sheet_name]
            grid = []
            for row in ws.iter_rows():
                out = []
                for cell in row:
                    v = cell.value
                    # Excel scales %-formatted cells by 100 for DISPLAY only (0.0577
                    # shown as "5.77%"). Narrative claims quote the displayed number,
                    # so store the display scale. Quoted literals ('0"%"') don't scale
                    # in Excel and are excluded. (.xls path: xlrd format lookup is
                    # impractical, so this correction is xlsx-only.)
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        fmt = re.sub(r'"[^"]*"', "", cell.number_format or "")
                        if "%" in fmt:
                            v = v * 100
                    out.append(v)
                grid.append(out)
            _apply_merged_fill(grid, [
                (m.min_row - 1, m.max_row, m.min_col - 1, m.max_col)
                for m in ws.merged_cells.ranges
            ])
            return grid
        finally:
            wb.close()
    else:
        raise ValueError("Unrecognized file format: expected .xls or .xlsx bytes.")


# ---------------------------------------------------------------------------
# Structure detection
# ---------------------------------------------------------------------------

def _is_empty(v) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def _is_number(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _find_header_row(grid: List[List]) -> int:
    for r in range(min(_MAX_HEADER_SCAN, len(grid))):
        nonempty = [v for v in grid[r] if not _is_empty(v)]
        if len(nonempty) < 2:
            continue
        headerish = [
            v for v in nonempty
            if (isinstance(v, str) and not _is_number(v)) or _parse_period(v) is not None
        ]
        if len(headerish) / len(nonempty) < _HEADERISH_MIN_FRAC:
            continue
        has_data_below = any(
            any(_is_number(v) for v in grid[rr])
            for rr in range(r + 1, min(r + 1 + _DATA_PROBE_ROWS, len(grid)))
        )
        if has_data_below:
            return r
    raise ValueError(
        "Tidak dapat mendeteksi baris header tabel: tidak ada baris teks dengan "
        "data numerik di bawahnya dalam 20 baris pertama."
    )


def _find_label_col(grid: List[List], header_row: int) -> int:
    n_cols = max((len(row) for row in grid), default=0)
    for c in range(n_cols):
        cells = [row[c] for row in grid[header_row + 1:] if c < len(row) and not _is_empty(row[c])]
        if not cells:
            continue
        text_cells = [v for v in cells if isinstance(v, str)]
        if len(text_cells) / len(cells) >= _LABEL_TEXT_MIN_FRAC:
            return c
    raise ValueError(
        "Tidak dapat mendeteksi kolom label: tidak ada kolom yang mayoritas berisi teks."
    )


def _title_and_unit(grid: List[List], header_row: int) -> Tuple[str, str]:
    title, unit = "", ""
    for r in range(header_row):
        row_texts = [str(v).strip() for v in grid[r] if not _is_empty(v)]
        if not row_texts:
            continue
        joined = " ".join(row_texts)
        m = re.fullmatch(r"\((.+)\)", joined)
        if m and not unit:
            unit = m.group(1).strip()
        elif not title:
            title = joined
    return title, unit


# ---------------------------------------------------------------------------
# Two-row header path: an int-year row (merged across period columns) directly above
# a row of bare quarter (I/II/III/IV, Q1..) or month tokens — the BI survey layout.
# Tried BEFORE the single-row heuristics because those latch onto the quarter row as a
# categorical header and silently produce garbage on such sheets. Regression safety
# comes from the strict structural preconditions below, not from ordering.
# ---------------------------------------------------------------------------

def _is_year_cell(v) -> bool:
    """True for an integral numeric cell in the plausible-year range (2012 or 2012.0)."""
    return (
        isinstance(v, (int, float)) and not isinstance(v, bool)
        and float(v).is_integer() and 1990 <= v <= 2100
    )


def _bare_period_token(v) -> Optional[str]:
    """Canonical token for a bare period header cell: 'Q1'..'Q4' or 'Jan'..'Dec'."""
    quarter = _parse_quarter_token(v)
    if quarter:
        return quarter
    if isinstance(v, str):
        return _MONTH_NAMES.get(v.strip().rstrip("*").strip().lower())
    return None


def _years_form_merged_runs(row: List, year_cols: List[int]) -> bool:
    """True when every distinct year value spans a contiguous run of >= 2 columns.

    Evidence the years came from horizontal merges (this layout), not the sparse
    single-cell anchors of the BI monthly layout — which tier 1 owns and whose year
    assignment would be wrong if naively read here.
    """
    runs: List[int] = []
    prev_c: Optional[int] = None
    for c in sorted(year_cols):
        if prev_c is not None and c == prev_c + 1 and row[c] == row[prev_c]:
            runs[-1] += 1
        else:
            runs.append(1)
        prev_c = c
    return bool(runs) and all(n >= 2 for n in runs)


def _find_header_blocks(grid: List[List]) -> List[Tuple[int, int]]:
    """All (year_row, period_row) header pairs, top to bottom, across the whole grid.

    Headers repeat mid-sheet when several tables are stacked in one sheet, so the scan
    is not limited to the top rows. A pair qualifies when the year row holds >= 2 year
    cells (merged years repeat after the fill) and >= _PERIOD_MIN_FRAC of those columns
    carry a parseable bare period token right below (computed ONLY over year-bearing
    columns — vertical label merges spill text into the period row).
    """
    blocks: List[Tuple[int, int]] = []
    r = 0
    while r < len(grid) - 1:
        year_cols = [c for c, v in enumerate(grid[r]) if _is_year_cell(v)]
        if len(year_cols) < 2:
            r += 1
            continue
        next_row = grid[r + 1]
        tokens = {
            c: tok for c in year_cols
            if (tok := _bare_period_token(next_row[c] if c < len(next_row) else None))
        }
        if len(tokens) < 2 or len(tokens) / len(year_cols) < _PERIOD_MIN_FRAC:
            r += 1
            continue
        if not any(t in _QUARTER_CANON for t in tokens.values()):
            # Month-only block: require merged-style contiguous year runs (see above).
            if not _years_form_merged_runs(grid[r], year_cols):
                r += 1
                continue
        blocks.append((r, r + 1))
        r += 2
    return blocks


def _block_col_keys(grid: List[List], year_row: int, period_row: int) -> Dict[int, Tuple[int, str]]:
    """{column -> (year, period_token)} for every column with a year above a token."""
    keys: Dict[int, Tuple[int, str]] = {}
    yrow, prow = grid[year_row], grid[period_row]
    for c, v in enumerate(yrow):
        if not _is_year_cell(v):
            continue
        tok = _bare_period_token(prow[c] if c < len(prow) else None)
        if tok:
            keys[c] = (int(v), tok)
    return keys


def _find_label_cols(
    grid: List[List], first_data_col: int, row_lo: int, row_hi: int
) -> List[int]:
    """Contiguous run of text-bearing columns ending right before the first data column."""
    def col_has_text(c: int) -> bool:
        for r in range(row_lo, min(row_hi, len(grid))):
            row = grid[r]
            if c < len(row) and isinstance(row[c], str) and row[c].strip():
                return True
        return False

    cols: List[int] = []
    c = first_data_col - 1
    while c >= 0 and col_has_text(c):
        cols.append(c)
        c -= 1
    return list(reversed(cols))


# Label-block cells that carry no name of their own: bullet glyphs and enumeration markers
# ('A.', 'B1.', '3)', '1)'). They must NOT overwrite the forward-filled section name of their
# own column. BI survey sheets put a section title ("Indeks Keyakinan Konsumen (IKK)") and the
# bullet of each row under it in the SAME column, so treating the bullet as a real label buried
# the section name: all nine sections of "Tabel 2" collapsed to 'A. > - > Pengeluaran >Rp5 juta',
# 'B. > - > Pengeluaran >Rp5 juta', … — indistinguishable to a claim naming the indicator.
_LABEL_MARKER_RE = re.compile(r"^(?:[\-–—•·*]+|(?:[A-Za-z]\d{0,2}|\d{1,2})[.)])$")


def _is_label_marker(v) -> bool:
    return isinstance(v, str) and bool(_LABEL_MARKER_RE.match(v.strip()))


def _composite_label(parts: List[Optional[str]]) -> str:
    """Join label parts with QUAL_SEP: whitespace-normalized, blanks dropped,
    consecutive case-insensitive duplicates collapsed ('Total > Total' -> 'Total')."""
    out: List[str] = []
    for p in parts:
        if p is None:
            continue
        text = re.sub(r"\s+", " ", str(p)).strip()
        if not text:
            continue
        if out and out[-1].lower() == text.lower():
            continue
        out.append(text)
    return QUAL_SEP.join(out)


def _drop_redundant_qualifiers(
    row_labels: List[str],
    table_data: Dict[Tuple, float],
    reduced: Dict[str, str],
) -> Tuple[List[str], Dict[Tuple, float]]:
    """Replace a label with its section-header-free form wherever the leaf is already unique.

    `reduced[label]` is the same label rebuilt without the levels that were INHERITED from a
    preceding row which held no data of its own — i.e. from a pure section-header row. Those
    are the only droppable levels: a level that came from this row's own cells (including a
    vertical merge, which fills every cell of the range) is part of the row's identity, so
    bilingual/parallel label blocks like 'Menurut Penggunaan > Kredit Modal Kerja > Based on
    Usage > Working Capital Loans' are never shortened.

    Dropping is then gated on the same convention as excel_parser_bi._qualify_duplicate_labels:
    a qualifier earns its place by resolving an ambiguity. It stays where it does — the
    expenditure-group sheet repeats 'Pengeluaran >Rp5 juta' under all nine indicator sections,
    and the per-city sheet repeats every metric name once per city. It goes where it does not:
    a sheet whose section header is ALSO one of its own rows ("A. Indeks Keyakinan Konsumen
    (IKK)" above the rows IKK / IKE / IEK) otherwise labels the national IKE row 'Indeks
    Keyakinan Konsumen (IKK) > … (IKE)', and those borrowed words make it look FARTHER from a
    plain "IKE" claim than a per-city or per-age-group IKE row is — so the national claim gets
    checked against one arbitrary breakdown row instead.
    """
    leaf_counts: Dict[str, int] = {}
    for label in row_labels:
        leaf = label.rsplit(QUAL_SEP, 1)[-1]
        leaf_counts[leaf] = leaf_counts.get(leaf, 0) + 1

    renames = {
        label: short
        for label in row_labels
        if (short := reduced.get(label, label)) != label
        and leaf_counts[label.rsplit(QUAL_SEP, 1)[-1]] == 1
    }
    if not renames:
        return row_labels, table_data

    new_labels = [renames.get(label, label) for label in row_labels]
    new_data = {
        (renames.get(key[0], key[0]), *key[1:]): value
        for key, value in table_data.items()
    }
    return new_labels, new_data


def _parse_two_row_table(grid: List[List]) -> Optional[TableData]:
    """Parse a sheet of stacked two-row-header tables into one temporal TableData.

    Each header block governs the rows down to the next block (or the sheet's end), so
    blank separator rows, repeated sub-headers, 'na' sentinels and footnote rows need no
    special casing: a row only enters the table when a mapped column holds a number.

    Row labels are composites of the label-block columns joined with QUAL_SEP, built
    with a HIERARCHICAL forward fill: a blank label cell inherits the value above it
    only while no column to its left is explicit in the same row, and an explicit value
    at level k clears the carried values of every level right of k — so a 'Total' row
    with a blank child column stays 'Total' instead of inheriting a stale child label.
    The parent qualifier is then DROPPED again wherever the leaf alone already identifies
    the row uniquely — see _drop_redundant_qualifiers.
    """
    blocks = _find_header_blocks(grid)
    if not blocks:
        return None

    table_data: Dict[Tuple, float] = {}
    row_labels: List[str] = []
    # label -> same label without the levels inherited from header-only rows; consumed by
    # _drop_redundant_qualifiers.
    reduced: Dict[str, str] = {}
    for i, (year_row, period_row) in enumerate(blocks):
        region_end = blocks[i + 1][0] if i + 1 < len(blocks) else len(grid)
        col_keys = _block_col_keys(grid, year_row, period_row)
        if not col_keys:
            continue
        label_cols = _find_label_cols(grid, min(col_keys), year_row, region_end)
        if not label_cols:
            continue

        carry: List[Optional[str]] = [None] * len(label_cols)
        # Per level: did the carried value come from a row that stored no data (a pure
        # section header) rather than from a data row's own cells?
        carry_from_header: List[bool] = [False] * len(label_cols)
        for r in range(period_row + 1, region_end):
            row = grid[r]
            parts: List[Optional[str]] = []
            own: List[bool] = []   # part belongs to this row's identity (not a header carry)
            explicit_left = False
            set_here: List[int] = []
            for idx, c in enumerate(label_cols):
                v = row[c] if c < len(row) else None
                # A bullet/enumeration marker counts as blank so the section name carried
                # down this column survives (see _is_label_marker).
                if not _is_empty(v) and not _is_label_marker(v):
                    carry[idx] = str(v)
                    for j in range(idx + 1, len(label_cols)):
                        carry[j] = None
                    parts.append(str(v))
                    own.append(True)
                    set_here.append(idx)
                    explicit_left = True
                else:
                    inherited = None if explicit_left else carry[idx]
                    parts.append(inherited)
                    own.append(inherited is None or not carry_from_header[idx])
            label = _composite_label(parts)
            if not label:
                # Nothing but markers on this row and nothing carried — fall back to the
                # raw cells so a marker-only key beats dropping the row's data entirely.
                label = _composite_label([
                    row[c] if c < len(row) else None for c in label_cols
                ])
                own = [True] * len(parts)
            if not label:
                continue
            stored = False
            for c, (year, tok) in col_keys.items():
                if c < len(row) and _is_number(row[c]):
                    table_data.setdefault((label, year, tok), float(row[c]))
                    stored = True
            # A row that stored nothing is a section header: what it set becomes droppable
            # context for the rows beneath it.
            for idx in set_here:
                carry_from_header[idx] = not stored
            if stored and label not in row_labels:
                row_labels.append(label)
                short = _composite_label([p for p, o in zip(parts, own) if o])
                reduced.setdefault(label, short or label)

    if not row_labels or not table_data:
        return None

    row_labels, table_data = _drop_redundant_qualifiers(row_labels, table_data, reduced)

    title, unit = _title_and_unit(grid, blocks[0][0])
    return TableData(
        title=title,
        unit=unit,
        row_labels=row_labels,
        col_labels=[],
        axis_type="temporal",
        _data=table_data,
    )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_generic_table(data: bytes, sheet_name: str) -> TableData:
    """Parse an arbitrary single-header-row table from .xls/.xlsx bytes into a TableData.

    Returns a TEMPORAL table (lookup by year/month) when the column headers read as
    calendar periods, otherwise a CATEGORICAL table (lookup by attribute/column name).
    Raises ValueError when no plausible table structure is found.
    """
    grid = _load_grid(data, sheet_name)

    # Two-row header layout (year row over quarter/month tokens) first: on such sheets
    # the single-row heuristics below misread the token row as a categorical header.
    two_row = _parse_two_row_table(grid)
    if two_row is not None:
        return two_row

    header_row = _find_header_row(grid)
    label_col = _find_label_col(grid, header_row)
    title, unit = _title_and_unit(grid, header_row)

    header = grid[header_row]
    data_cols = [
        c for c in range(label_col + 1, len(header)) if not _is_empty(header[c])
    ]
    if not data_cols:
        raise ValueError("Tidak ada kolom data di kanan kolom label.")

    col_periods = {c: _parse_period(header[c]) for c in data_cols}
    n_period_cols = sum(1 for p in col_periods.values() if p is not None)
    temporal = n_period_cols >= 2 and n_period_cols / len(data_cols) >= _PERIOD_MIN_FRAC

    col_keys: Dict[int, Tuple] = {}
    col_labels: List[str] = []
    if temporal:
        col_keys = {c: p for c, p in col_periods.items() if p is not None}
    else:
        for c in data_cols:
            name = str(header[c]).strip()
            if name and name not in col_labels:
                col_labels.append(name)
                col_keys[c] = (name,)

    table_data: Dict[Tuple, float] = {}
    row_labels: List[str] = []
    for row in grid[header_row + 1:]:
        if label_col >= len(row) or _is_empty(row[label_col]):
            continue
        label = str(row[label_col]).strip()
        stored = False
        for c, key_part in col_keys.items():
            if c < len(row) and _is_number(row[c]):
                table_data.setdefault((label, *key_part), float(row[c]))
                stored = True
        if stored and label not in row_labels:
            row_labels.append(label)

    if not row_labels or not table_data:
        raise ValueError(
            "Tabel terdeteksi tetapi tidak ada baris berlabel dengan nilai numerik."
        )

    return TableData(
        title=title,
        unit=unit,
        row_labels=row_labels,
        col_labels=col_labels,
        axis_type="temporal" if temporal else "categorical",
        _data=table_data,
    )
