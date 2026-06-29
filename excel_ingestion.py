"""Ingests stylized Excel statistical tables into the long/tidy `excel_facts` SQL table.

Excel files are read via `openpyxl` (not pandas/CSV) because cell *style* - bold, fill, borders,
merged ranges - carries structural meaning in these source tables (which cells are headers, which
are totals) that plain values alone don't capture. Header bands, row/column labels, and aggregate
rows/columns are detected with style+keyword+position heuristics that produce a confidence score
per row/column; only rows/columns in the resulting "ambiguous" confidence band are escalated to an
LLM call, which labels roles from text and pre-computed boolean signals only - it never sees or
repeats numeric values, so it cannot hallucinate a number into the ground-truth dataset.

Regardless of the original file's column order, header offset, or row/column orientation of its
aggregate cells, every sheet is normalized into the same long/tidy shape (source_file, sheet,
row_label, col_label, value, is_aggregate, source_cell_ref) before being written to SQLite - so
`db.py`/`verifier.py` see one stable schema no matter how many quirky source files get ingested.
"""

import io
import logging
import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Literal, Optional, Tuple

from langchain_core.language_models import BaseChatModel
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.runnables import Runnable
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field

from db import DB_PATH as DEFAULT_DB_PATH

logger = logging.getLogger("fact-checker")

# Indonesian/English keywords that mark a row or column label as a derived total/average rather
# than an independent data point.
AGGREGATE_KEYWORDS = re.compile(
    r"jumlah|grand\s*total|sub\s*-?\s*total|\btotal\b|rata-?rata|rata\s*2|average|\bmean\b",
    re.IGNORECASE,
)

# A row/column with score >= CONFIDENT_AGGREGATE is auto-flagged as an aggregate; below
# CONFIDENT_NOT_AGGREGATE it's auto-flagged as plain data; in between, it's ambiguous and gets
# escalated to the LLM (or defaults to "not aggregate" if no LLM is available).
CONFIDENT_AGGREGATE = 0.7
CONFIDENT_NOT_AGGREGATE = 0.3

FACTS_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS excel_facts (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    source_file     TEXT NOT NULL,
    sheet           TEXT NOT NULL,
    row_label       TEXT NOT NULL,
    col_label       TEXT NOT NULL,
    value           REAL NOT NULL,
    is_aggregate    INTEGER NOT NULL,
    source_cell_ref TEXT NOT NULL
);
"""


@dataclass
class CellInfo:
    """A cell's value plus the style attributes used as structural signals."""
    value: object
    is_bold: bool
    has_fill: bool
    has_top_border: bool


CellGrid = Dict[Tuple[int, int], CellInfo]


@dataclass
class AggregateSignal:
    """The pre-computed features behind one row/column's aggregate-confidence score."""
    index: int
    label: str
    keyword_match: bool
    style_signal: bool
    numeric_consistency: bool
    score: float


@dataclass
class IngestSummary:
    n_sheets: int = 0
    n_facts: int = 0
    auto_aggregate: int = 0
    auto_not_aggregate: int = 0
    llm_escalated: int = 0
    defaulted: int = 0

    def add_sheet_counts(self, counts: Dict[str, int]) -> None:
        if not counts:
            return
        self.n_sheets += 1
        self.auto_aggregate += counts.get("auto_aggregate", 0)
        self.auto_not_aggregate += counts.get("auto_not_aggregate", 0)
        self.llm_escalated += counts.get("llm_escalated", 0)
        self.defaulted += counts.get("defaulted", 0)


# --- Reading cells with style, including merged-cell propagation -----------------------------

def _read_sheet_cells(ws: Worksheet) -> CellGrid:
    """Read every non-empty cell's value+style, propagating merged-range anchors to their span.

    openpyxl only stores the value/style on a merged range's top-left cell; every other cell in
    the range comes back from `iter_rows()` as an empty `MergedCell`. Propagating the anchor's
    value/style to the rest of the range means downstream logic never has to special-case merges.
    """
    cells: CellGrid = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            cells[(c.row, c.column)] = CellInfo(
                value=c.value,
                is_bold=bool(c.font and c.font.bold),
                has_fill=bool(c.fill and c.fill.patternType is not None),
                has_top_border=bool(c.border and c.border.top and c.border.top.style is not None),
            )

    for merged_range in ws.merged_cells.ranges:
        anchor = cells.get((merged_range.min_row, merged_range.min_col))
        if anchor is None:
            continue
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if (r, col) == (merged_range.min_row, merged_range.min_col):
                    continue
                cells[(r, col)] = anchor

    return cells


# --- Header band / data band / label vs value column detection -------------------------------

def _is_title_row(row_cells: Dict[int, CellInfo], min_col: int, max_col: int) -> bool:
    """A decorative title row: after merge-propagation, the whole row repeats one single value
    spanning at least half the table's width (e.g. a merged "Laporan Penjualan 2024" banner)."""
    if len(row_cells) < 2:
        return False
    distinct_values = {cell.value for cell in row_cells.values()}
    span = max_col - min_col + 1
    return len(distinct_values) == 1 and len(row_cells) >= max(2, int(span * 0.5))


def _is_header_like_row(row_cells: Dict[int, CellInfo]) -> bool:
    """A row belongs to the header band if most of its cells are text.

    Limitation: a header row made entirely of bare numbers (e.g. year columns typed as 2023/2024
    rather than text) won't be caught by this rule and would be misread as the start of the data
    band. A bold/fill-based fallback could catch that case, but it would also have to be guarded
    against colliding with a fully-bold aggregate/total row positioned right after the header band
    with no detail rows in between - not handled here; flagged as a known gap for this PoC.
    """
    non_empty = [cell for cell in row_cells.values() if cell.value is not None]
    if not non_empty:
        return False
    text_count = sum(1 for cell in non_empty if isinstance(cell.value, str))
    # Strictly greater than half (not >=): a row with exactly one text label column and one
    # numeric value column - very common in narrow (2-3 column) tables - sits at exactly 0.5 and
    # must NOT be misread as a header, or every data row in a 2-column table would qualify.
    return text_count / len(non_empty) > 0.5


def _split_header_and_data_rows(
    cells: CellGrid, rows_with_content: List[int], min_col: int, max_col: int
) -> Tuple[List[int], List[int]]:
    """Walk rows top-down: skip leading decorative title row(s), then classify a leading run of
    mostly-text/bold rows as the header band, and everything after that as the data band.

    This is a simplification: it assumes one contiguous header band followed by one contiguous
    data band with no trailing footnote section - good enough for a single statistical table per
    sheet, but a sheet with multiple stacked tables would need this split run per sub-table.
    """
    header_rows: List[int] = []
    data_rows: List[int] = []
    started = False
    in_header = True
    for r in rows_with_content:
        row_cells = {c: cell for (rr, c), cell in cells.items() if rr == r}
        if not row_cells:
            continue
        if not started and _is_title_row(row_cells, min_col, max_col):
            continue
        started = True
        if in_header and _is_header_like_row(row_cells):
            header_rows.append(r)
        else:
            in_header = False
            data_rows.append(r)
    return header_rows, data_rows


def _split_label_and_value_cols(
    cells: CellGrid, data_rows: List[int], min_col: int, max_col: int
) -> Tuple[List[int], List[int]]:
    """Within the data band, a column is a value column if most of its cells are numeric, and a
    row-label column otherwise (e.g. a "Produk"/"Cabang" name column)."""
    label_cols: List[int] = []
    value_cols: List[int] = []
    for c in range(min_col, max_col + 1):
        col_cells = [cells[(r, c)] for r in data_rows if (r, c) in cells and cells[(r, c)].value is not None]
        if not col_cells:
            continue
        numeric_count = sum(1 for cell in col_cells if isinstance(cell.value, (int, float)))
        if numeric_count / len(col_cells) >= 0.5:
            value_cols.append(c)
        else:
            label_cols.append(c)
    return label_cols, value_cols


def _compute_header_paths(cells: CellGrid, header_rows: List[int], value_cols: List[int]) -> Dict[int, str]:
    """Build each value column's multi-level header path, e.g. "Semester 1 > Q1"."""
    paths: Dict[int, str] = {}
    for c in value_cols:
        parts = []
        for r in header_rows:
            cell = cells.get((r, c))
            if cell is not None and cell.value is not None and str(cell.value) not in parts:
                parts.append(str(cell.value))
        paths[c] = " > ".join(parts) if parts else get_column_letter(c)
    return paths


def _row_label(cells: CellGrid, row: int, label_cols: List[int]) -> str:
    parts = [str(cells[(row, c)].value) for c in label_cols if (row, c) in cells and cells[(row, c)].value is not None]
    return " - ".join(parts)


# --- Aggregate row/column detection -----------------------------------------------------------

def _row_matches_sum_of_others(cells: CellGrid, row: int, other_rows: List[int], value_cols: List[int]) -> bool:
    if len(other_rows) < 2:
        return False
    checked = matched = 0
    for c in value_cols:
        this_cell = cells.get((row, c))
        if this_cell is None or not isinstance(this_cell.value, (int, float)):
            continue
        total = 0.0
        for r2 in other_rows:
            other = cells.get((r2, c))
            if other is None or not isinstance(other.value, (int, float)):
                total = None
                break
            total += other.value
        if total is None:
            continue
        checked += 1
        if abs(total - this_cell.value) <= max(0.01, abs(this_cell.value) * 0.005):
            matched += 1
    return checked > 0 and matched == checked


def _col_matches_sum_of_others(cells: CellGrid, col: int, other_cols: List[int], data_rows: List[int]) -> bool:
    if len(other_cols) < 2:
        return False
    checked = matched = 0
    for r in data_rows:
        this_cell = cells.get((r, col))
        if this_cell is None or not isinstance(this_cell.value, (int, float)):
            continue
        total = 0.0
        for c2 in other_cols:
            other = cells.get((r, c2))
            if other is None or not isinstance(other.value, (int, float)):
                total = None
                break
            total += other.value
        if total is None:
            continue
        checked += 1
        if abs(total - this_cell.value) <= max(0.01, abs(this_cell.value) * 0.005):
            matched += 1
    return checked > 0 and matched == checked


def _score_row_aggregate(
    cells: CellGrid, row: int, data_rows: List[int], label_cols: List[int], value_cols: List[int]
) -> AggregateSignal:
    label = _row_label(cells, row, label_cols)
    keyword_match = bool(AGGREGATE_KEYWORDS.search(label))
    value_cells = [cells[(row, c)] for c in value_cols if (row, c) in cells]
    style_signal = bool(value_cells) and (
        all(vc.is_bold for vc in value_cells)
        or all(vc.has_top_border for vc in value_cells)
        or all(vc.has_fill for vc in value_cells)
    )
    numeric_consistency = _row_matches_sum_of_others(cells, row, [r for r in data_rows if r != row], value_cols)

    score = 0.0
    if keyword_match:
        score += 0.5
    if style_signal:
        score += 0.2
    if numeric_consistency:
        score += 0.6
    if data_rows and row == max(data_rows):
        score += 0.05
    return AggregateSignal(row, label, keyword_match, style_signal, numeric_consistency, min(score, 1.0))


def _score_col_aggregate(
    cells: CellGrid, col: int, data_rows: List[int], value_cols: List[int], header_paths: Dict[int, str]
) -> AggregateSignal:
    label = header_paths.get(col, "")
    keyword_match = bool(AGGREGATE_KEYWORDS.search(label))
    numeric_consistency = _col_matches_sum_of_others(cells, col, [c for c in value_cols if c != col], data_rows)

    score = 0.0
    if keyword_match:
        score += 0.5
    if numeric_consistency:
        score += 0.6
    if value_cols and col == max(value_cols):
        score += 0.05
    # No style signal for columns yet: header cells are typically styled uniformly regardless of
    # which one is the total, unlike aggregate rows where the whole row is usually styled apart.
    return AggregateSignal(col, label, keyword_match, False, numeric_consistency, min(score, 1.0))


def _classify(score: float) -> Literal["aggregate", "not_aggregate", "ambiguous"]:
    if score >= CONFIDENT_AGGREGATE:
        return "aggregate"
    if score < CONFIDENT_NOT_AGGREGATE:
        return "not_aggregate"
    return "ambiguous"


# --- LLM escalation for ambiguous rows/columns ------------------------------------------------
# The LLM only ever sees text labels and pre-computed booleans here - never a raw cell value -
# so it cannot introduce a hallucinated number into the ground-truth dataset; it only decides a
# role (aggregate vs. data), and the actual values are still pulled by code from their original
# coordinates regardless of what the LLM decides.

AGGREGATE_LABEL_SYSTEM_PROMPT = """You are classifying rows or columns of a statistical table as either
AGGREGATE (a total, subtotal, or average derived from the other rows/columns) or DATA (an independent
observation that must not be treated as a derived total).

You are given each ambiguous item's text label plus three pre-computed signals - you are never given
raw numeric values:
- keyword_match: the label contains a word like "jumlah"/"total"/"rata-rata".
- style_signal: the row/column was styled distinctly from the rest (e.g. bold, or a separating border).
- numeric_consistency: the row/column's values are numerically consistent with being the sum of the
  other rows/columns in the same table.

Weigh the label text and all three signals together - no single signal is decisive alone. Return
exactly one verdict per item, tagged with its index.
"""

AGGREGATE_LABEL_PROMPT = ChatPromptTemplate.from_messages(
    [
        ("system", AGGREGATE_LABEL_SYSTEM_PROMPT),
        ("human", "Ambiguous items:\n{items_block}\n\nLabel each item as aggregate or data."),
    ]
)


class IndexedAggregateLabel(BaseModel):
    index: int = Field(..., description="Index of the item this verdict is for, matching the input list.")
    is_aggregate: bool = Field(..., description="True if this row/column is a derived total/average, not independent data.")
    reasoning: str = Field(..., description="Concise explanation referencing the label and the three signals.")


class BatchAggregateLabels(BaseModel):
    labels: List[IndexedAggregateLabel] = Field(..., description="One verdict per ambiguous item.")


def build_aggregate_label_chain(llm: BaseChatModel) -> Runnable:
    return AGGREGATE_LABEL_PROMPT | llm.with_structured_output(BatchAggregateLabels)


def _label_ambiguous_with_llm(items: List[AggregateSignal], llm: BaseChatModel) -> Dict[int, bool]:
    chain = build_aggregate_label_chain(llm)
    items_block = "\n".join(
        f"{it.index}. label={it.label!r}, keyword_match={it.keyword_match}, "
        f"style_signal={it.style_signal}, numeric_consistency={it.numeric_consistency}"
        for it in items
    )
    try:
        response: BatchAggregateLabels = chain.invoke({"items_block": items_block})
    except Exception:
        logger.exception("Ambiguous aggregate labeling failed; defaulting these items to is_aggregate=False")
        return {}
    return {label.index: label.is_aggregate for label in response.labels}


def _resolve_scores(signals: List[AggregateSignal], llm: Optional[BaseChatModel], kind: str) -> Tuple[Dict[int, bool], Dict[str, int]]:
    result: Dict[int, bool] = {}
    counts = {"auto_aggregate": 0, "auto_not_aggregate": 0, "llm_escalated": 0, "defaulted": 0}
    ambiguous = []
    for sig in signals:
        verdict = _classify(sig.score)
        if verdict == "aggregate":
            result[sig.index] = True
            counts["auto_aggregate"] += 1
        elif verdict == "not_aggregate":
            result[sig.index] = False
            counts["auto_not_aggregate"] += 1
        else:
            ambiguous.append(sig)

    if ambiguous:
        if llm is not None:
            labeled = _label_ambiguous_with_llm(ambiguous, llm)
            for sig in ambiguous:
                result[sig.index] = labeled.get(sig.index, False)
            counts["llm_escalated"] += len(ambiguous)
        else:
            logger.warning("%d ambiguous %s(s) defaulted to is_aggregate=False (no LLM provided)", len(ambiguous), kind)
            for sig in ambiguous:
                result[sig.index] = False
            counts["defaulted"] += len(ambiguous)

    return result, counts


# --- Per-sheet extraction into long/tidy records -----------------------------------------------

def extract_sheet_facts(
    filename: str, sheet_name: str, ws: Worksheet, llm: Optional[BaseChatModel] = None
) -> Tuple[List[dict], Dict[str, int]]:
    """Normalize one sheet into long/tidy fact records, regardless of its original column order,
    header offset, or whether its aggregate cell is a row, a column, or absent altogether."""
    cells = _read_sheet_cells(ws)
    if not cells:
        return [], {}

    cols_seen = [c for (_, c) in cells]
    min_col, max_col = min(cols_seen), max(cols_seen)
    rows_with_content = sorted({r for (r, _) in cells})

    header_rows, data_rows = _split_header_and_data_rows(cells, rows_with_content, min_col, max_col)
    if not data_rows:
        return [], {}

    label_cols, value_cols = _split_label_and_value_cols(cells, data_rows, min_col, max_col)
    if not value_cols:
        return [], {}

    header_paths = _compute_header_paths(cells, header_rows, value_cols)

    row_signals = [_score_row_aggregate(cells, r, data_rows, label_cols, value_cols) for r in data_rows]
    col_signals = [_score_col_aggregate(cells, c, data_rows, value_cols, header_paths) for c in value_cols]

    row_aggregate, row_counts = _resolve_scores(row_signals, llm, "row")
    col_aggregate, col_counts = _resolve_scores(col_signals, llm, "column")

    records = []
    for r in data_rows:
        row_label = _row_label(cells, r, label_cols)
        for c in value_cols:
            cell = cells.get((r, c))
            if cell is None or not isinstance(cell.value, (int, float)):
                continue
            records.append(
                {
                    "source_file": filename,
                    "sheet": sheet_name,
                    "row_label": row_label,
                    "col_label": header_paths.get(c, get_column_letter(c)),
                    "value": float(cell.value),
                    "is_aggregate": bool(row_aggregate.get(r, False) or col_aggregate.get(c, False)),
                    "source_cell_ref": f"{sheet_name}!{get_column_letter(c)}{r}",
                }
            )

    combined_counts = {
        key: row_counts.get(key, 0) + col_counts.get(key, 0)
        for key in ("auto_aggregate", "auto_not_aggregate", "llm_escalated", "defaulted")
    }
    return records, combined_counts


# --- Writing facts to SQLite --------------------------------------------------------------------

def write_facts(records: List[dict], db_path: Path) -> None:
    """Write facts via a direct (writable) sqlite3 connection - `db.py`'s read-only connection is
    for the app/verifier at query time, ingestion is a separate offline step, same pattern as
    `setup_db.py`. Re-ingesting a file deletes its previous rows first, so reruns are idempotent.
    """
    if not records:
        return
    conn = sqlite3.connect(db_path)
    try:
        cursor = conn.cursor()
        cursor.execute(FACTS_TABLE_SQL)
        for source_file in {r["source_file"] for r in records}:
            cursor.execute("DELETE FROM excel_facts WHERE source_file = ?", (source_file,))
        cursor.executemany(
            """
            INSERT INTO excel_facts
                (source_file, sheet, row_label, col_label, value, is_aggregate, source_cell_ref)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    r["source_file"], r["sheet"], r["row_label"], r["col_label"],
                    r["value"], int(r["is_aggregate"]), r["source_cell_ref"],
                )
                for r in records
            ],
        )
        conn.commit()
    finally:
        conn.close()


def _ingest_workbook(wb, filename: str, db_path: Optional[Path], llm: Optional[BaseChatModel]) -> IngestSummary:
    summary = IngestSummary()
    all_records: List[dict] = []

    for sheet_name in wb.sheetnames:
        records, counts = extract_sheet_facts(filename, sheet_name, wb[sheet_name], llm)
        all_records.extend(records)
        summary.add_sheet_counts(counts)

    write_facts(all_records, db_path or DEFAULT_DB_PATH)
    summary.n_facts = len(all_records)
    return summary


def ingest_file(path: Path, db_path: Optional[Path] = None, llm: Optional[BaseChatModel] = None) -> IngestSummary:
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    return _ingest_workbook(wb, path.name, db_path, llm)


def ingest_bytes(
    file_bytes: bytes, filename: str, db_path: Optional[Path] = None, llm: Optional[BaseChatModel] = None
) -> IngestSummary:
    """Same as `ingest_file`, but for an in-memory upload (e.g. a FastAPI `UploadFile`) - no temp
    file needed, `openpyxl` reads directly from a `BytesIO`, mirroring `pdf_extraction.py`'s
    bytes-in/no-disk-I/O convention for uploaded documents."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    return _ingest_workbook(wb, filename, db_path, llm)


if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO)
    paths = sys.argv[1:]
    if not paths:
        print("Usage: python excel_ingestion.py <file.xlsx> [<file2.xlsx> ...]")
        raise SystemExit(1)

    llm_for_escalation: Optional[BaseChatModel] = None
    try:
        from llm_provider import get_llm

        llm_for_escalation = get_llm(temperature=0.0)
    except Exception as exc:
        print(
            f"Warning: no LLM available for ambiguous-cell labeling ({exc}); "
            "ambiguous rows/columns will default to is_aggregate=False."
        )

    for file_path in paths:
        result = ingest_file(file_path, llm=llm_for_escalation)
        print(
            f"{file_path}: {result.n_facts} facts from {result.n_sheets} sheet(s) - "
            f"{result.auto_aggregate} auto-flagged aggregate, {result.auto_not_aggregate} auto-flagged data, "
            f"{result.llm_escalated} resolved via LLM, {result.defaulted} defaulted without LLM."
        )
